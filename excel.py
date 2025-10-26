import os
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Optional, Any, Tuple
from jsonhandler import JsonHandler
from utility import Utilities
from transactions import Transaction


class ExcelHandler:
    """Handler for Excel import and export operations"""
    
    REQUIRED_COLUMNS = ["Type", "Amount", "Category", "Date", "Description", "Payment Method"]
    
    # Valid values for each column
    VALID_TYPES = ["expense", "income"]
    VALID_PAYMENT_METHODS = ["Cash", "Credit Card", "Debit Card", "Bank Transfer"]
    VALID_EXPENSE_CATEGORIES = ["Food", "Transport", "Bills", "Shopping", "Entertainment", "Other"]
    VALID_INCOME_CATEGORIES = ["Salary", "Freelance", "Investment", "Gift", "Other"]
    
    def __init__(self, user_id: str):
        self.user_id = user_id
        self.json_handler = JsonHandler()
        self.utilities = Utilities()
    
    def validate_format(self, file_path: str) -> Tuple[bool, str]:
        """Validate the Excel file format
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            if not os.path.exists(file_path):
                return False, "File not found"
            
            # Load workbook
            workbook = load_workbook(file_path)
            
            if not workbook.active:
                return False, "Excel file is empty"
            
            sheet = workbook.active
            
            # Check if we have data
            if sheet.max_row < 2:
                return False, "Excel file must have at least one data row (in addition to header)"
            
            # Get header row
            headers = [cell.value for cell in sheet[1]]
            
            # Check if all required columns exist
            missing_columns = []
            for required_col in self.REQUIRED_COLUMNS:
                if required_col not in headers:
                    missing_columns.append(required_col)
            
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
            
            # Validate data rows
            errors = []
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=row_num, column=col_idx)
                    row_data[header] = cell.value
                
                # Validate this row
                row_error = self._validate_row(row_data, row_num)
                if row_error:
                    errors.append(row_error)
            
            if errors:
                return False, "\n".join(errors[:10])  # Return first 10 errors
            
            return True, ""
            
        except Exception as e:
            return False, f"Error validating file: {str(e)}"
    
    def _validate_row(self, row_data: Dict[str, Any], row_num: int) -> Optional[str]:
        """Validate a single row of data"""
        # Check required fields
        if not row_data.get("Type"):
            return f"Row {row_num}: Type is required"
        if not row_data.get("Amount"):
            return f"Row {row_num}: Amount is required"
        if not row_data.get("Category"):
            return f"Row {row_num}: Category is required"
        if not row_data.get("Date"):
            return f"Row {row_num}: Date is required"
        
        # Validate Type
        type_value = str(row_data["Type"]).strip().lower()
        if type_value not in self.VALID_TYPES:
            return f"Row {row_num}: Type must be one of {', '.join(self.VALID_TYPES)}"
        
        # Validate Amount
        try:
            amount = float(row_data["Amount"])
            if amount <= 0:
                return f"Row {row_num}: Amount must be positive"
        except (ValueError, TypeError):
            return f"Row {row_num}: Amount must be a valid number"
        
        # Validate Category
        category = str(row_data["Category"]).strip()
        if type_value == "expense":
            if category not in self.VALID_EXPENSE_CATEGORIES:
                return f"Row {row_num}: Category for expense must be one of {', '.join(self.VALID_EXPENSE_CATEGORIES)}"
        else:
            if category not in self.VALID_INCOME_CATEGORIES:
                return f"Row {row_num}: Category for income must be one of {', '.join(self.VALID_INCOME_CATEGORIES)}"
        
        # Validate Date
        date_value = str(row_data["Date"])
        if not self.utilities.validate_date(date_value):
            return f"Row {row_num}: Date must be in YYYY-MM-DD format"
        
        # Validate Payment Method (if provided)
        payment_method = row_data.get("Payment Method", "")
        if payment_method and payment_method not in self.VALID_PAYMENT_METHODS:
            return f"Row {row_num}: Payment Method must be one of {', '.join(self.VALID_PAYMENT_METHODS)}"
        
        return None
    
    def import_excel_transactions(self, file_path: str) -> Tuple[bool, str, int]:
        """Import transactions from Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple of (success, message, imported_count)
        """
        try:
            # Validate file format first
            is_valid, error_msg = self.validate_format(file_path)
            if not is_valid:
                return False, error_msg, 0
            
            # Load workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            # Get existing transactions
            transactions_data = self.json_handler.load_transactions()
            
            imported_count = 0
            failed_count = 0
            
            # Process each row
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=row_num, column=col_idx)
                    row_data[header] = cell.value
                
                try:
                    # Create transaction using the TransactionManager
                    transaction = Transaction(
                        user_id=self.user_id,
                        type=str(row_data["Type"]).strip().lower(),
                        amount=float(row_data["Amount"]),
                        category=str(row_data["Category"]).strip(),
                        date=str(row_data["Date"]),
                        description=str(row_data.get("Description", "") or ""),
                        payment_method=str(row_data.get("Payment Method", "Cash") or "Cash")
                    )
                    
                    # Add transaction
                    result = transaction.add_transaction()
                    if result:
                        imported_count += 1
                    else:
                        failed_count += 1
                        
                except Exception as e:
                    failed_count += 1
                    continue
            
            message = f"Import completed: {imported_count} transactions imported"
            if failed_count > 0:
                message += f", {failed_count} failed"
            
            return True, message, imported_count
            
        except Exception as e:
            return False, f"Error importing transactions: {str(e)}", 0
    
    def export_excel_transactions(self, file_path: Optional[str] = None) -> Tuple[bool, str]:
        """Export transactions to Excel file
        
        Args:
            file_path: Optional custom path for the Excel file
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Load transactions
            transactions_data = self.json_handler.load_transactions()
            user_transactions = transactions_data.get(self.user_id, [])
            
            if not user_transactions:
                return False, "No transactions found to export"
            
            # If no file path provided, generate one
            if not file_path:
                file_path = f"transactions_export_{self.user_id[:8]}.xlsx"
            
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Transactions"
            
            # Write headers
            headers = ["Transaction ID", "Type", "Amount", "Category", "Date", 
                      "Description", "Payment Method"]
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", 
                                                         end_color="CCCCCC", 
                                                         fill_type="solid")
            
            # Write data
            for row_idx, transaction in enumerate(user_transactions, start=2):
                sheet.cell(row=row_idx, column=1, value=transaction.get("transaction_id"))
                sheet.cell(row=row_idx, column=2, value=transaction.get("type"))
                sheet.cell(row=row_idx, column=3, value=transaction.get("amount"))
                sheet.cell(row=row_idx, column=4, value=transaction.get("category"))
                sheet.cell(row=row_idx, column=5, value=transaction.get("date"))
                sheet.cell(row=row_idx, column=6, value=transaction.get("description"))
                sheet.cell(row=row_idx, column=7, value=transaction.get("payment_method"))
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, start=1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = 15
            
            # Save workbook
            workbook.save(file_path)
            
            return True, f"Transactions exported successfully to {file_path}"
            
        except Exception as e:
            return False, f"Error exporting transactions: {str(e)}"

